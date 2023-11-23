import os
import time
import psutil
from openpyxl import load_workbook
from tqdm import tqdm

def main():
    # Caminhos
    userprofile = os.environ['USERPROFILE']
    filesPath = os.path.join(userprofile, "Documentos", "PythonVsGo_Dataset", "documentos_excel")
    logPath = os.path.join(userprofile, "Área de Trabalho", "Projetos", "PythonVsGo-FileManipulation", "PythonVsGo-FileManipulation", "logs", "unique_python_log.txt")

    # Monitoramento de desempenho (início)
    startTime = time.time()
    cpuUsageBefore = psutil.cpu_percent()
    memUsageBefore = psutil.virtual_memory().percent

    # Processamento dos arquivos
    uniqueLinesCount, err = countUniqueLines(filesPath)
    if err is not None:
        print("Erro ao processar arquivos:", err)
        return

    # Monitoramento de desempenho (fim)
    endTime = time.time()
    elapsedTime = endTime - startTime
    cpuUsageAfter = psutil.cpu_percent()
    memUsageAfter = psutil.virtual_memory().percent

    # Salvando em um arquivo de log
    logData = f"Time elapsed: {elapsedTime} seconds\n"
    logData += f"CPU Usage Before: {cpuUsageBefore}%, After: {cpuUsageAfter}%\n"
    logData += f"Memory Usage Before: {memUsageBefore}%, After: {memUsageAfter}%\n"

    for file, count in uniqueLinesCount.items():
        logData += f"{file}: {count} linhas únicas\n"

    try:
        with open(logPath, "w") as logFile:
            logFile.write(logData)
    except Exception as e:
        print("Erro ao escrever no arquivo de log:", e)
        return

    print("Process finished!")

# countUniqueLines conta linhas únicas em arquivos Excel no diretório especificado
def countUniqueLines(dir):
    uniqueLines = {}
    files = [f for f in os.listdir(dir) if f.endswith(".xlsx")]
    with tqdm(total=len(files), desc="Processando arquivos") as pbar:
        for file in files:
            file_path = os.path.join(dir, file)
            uniqueRowCount, err = countUniqueRows(file_path)
            if err is None:
                uniqueLines[file] = uniqueRowCount
            pbar.update(1)
    return uniqueLines, None

# countUniqueRows conta as linhas únicas em um arquivo Excel
def countUniqueRows(file_path):
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        uniqueRows = set()
        for row in sheet.iter_rows(values_only=True):
            uniqueRows.add(tuple(row))
        return len(uniqueRows), None
    except Exception as e:
        return None, e

if __name__ == "__main__":
    main()
