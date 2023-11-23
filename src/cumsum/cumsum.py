import os
import sys
import time
from openpyxl import load_workbook
from tqdm import tqdm
import psutil

def main():
    userprofile = os.environ['USERPROFILE']
    files_path = os.path.join(userprofile, "Documentos", "PythonVsGo_Dataset", "documentos_excel")
    log_path = os.path.join(userprofile, "Área de Trabalho", "Projetos", "PythonVsGo-FileManipulation", "PythonVsGo-FileManipulation", "logs")

    start_time = time.time()
    cpu_usage_before = psutil.cpu_percent()
    memory_usage_before = psutil.virtual_memory().percent

    cumsum_results, err = calculate_cumsum(files_path)
    if err is not None:
        print("Erro ao processar arquivos:", err)
        return

    cpu_usage_after = psutil.cpu_percent()
    memory_usage_after = psutil.virtual_memory().percent
    end_time = time.time()
    elapsed_time = end_time - start_time

    with open(os.path.join(log_path, "cumsum_python_log.txt"), "w") as log_file:
        log_file.write(f"Time elapsed: {elapsed_time} seconds\n")
        log_file.write(f"CPU Usage Before: {cpu_usage_before}%, After: {cpu_usage_after}%\n")
        log_file.write(f"Memory Usage Before: {memory_usage_before}%, After: {memory_usage_after}%\n")
        for file, result in cumsum_results.items():
            log_file.write(f"{file}: Cumsum = {result}\n")

    print("Process finished!")

def calculate_cumsum(directory):
    results = {}
    files = os.listdir(directory)
    with tqdm(total=len(files), desc="Processando arquivos") as pbar:
        for file in files:
            if file.endswith(".xlsx"):
                file_path = os.path.join(directory, file)
                cumsum, err = calculate_file_cumsum(file_path)
                if err is None:
                    results[file] = cumsum
                pbar.update(1)
    return results, None

def calculate_file_cumsum(file_path):
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        cumsum = 0.0
        for row in sheet.iter_rows(values_only=True):
            if row and isinstance(row[0], (int, float)):
                cumsum += float(row[0])
        return cumsum, None
    except Exception as e:
        return None, e

if __name__ == "__main__":
    main()
